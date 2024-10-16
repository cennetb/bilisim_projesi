import cv2
import numpy as np
import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook

def image_show(image, window_name):
    """Verilen görüntüyü belirtilen pencere adıyla ekranda gösterir."""
    cv2.imshow(window_name, image)

def gri(image):
    """Görüntü üzerinde bir işlem yap ve yeni bir görüntü döndür."""
    # Örnek işlem: Görüntüyü gri tonlamaya çevir
    return cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

def iso(image):
    """Görüntüyü ikili hale getirme (Isodata thresholding)."""
    _, isodata_goruntu = cv2.threshold(image, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    return isodata_goruntu

def apply_median_filter(image, kernel_size):
    """Median filtreleme uygula."""
    filtered_image = cv2.medianBlur(image, kernel_size)
    return filtered_image

def find_contours(image):
    """Görüntüdeki şekillerin sınırlarını bulur ve koordinatlarını döndürür."""
    if len(image.shape) == 3:
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    else:
        gray = image
    _, binary = cv2.threshold(gray, 127, 255, cv2.THRESH_BINARY_INV)
    contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    return contours

def draw_contours(image, contours):
    """Görüntüdeki şekillerin sınırlarını çiz ve ekranda göster.--pixel ayarlama"""
    cv2.drawContours(image, contours, -1, (0, 0, 0), 1)


def draw_circles(image, contours):
    """Görüntüdeki şekillerin etrafına minimum çemberler çizer ve çaplarını hesaplar."""
    circle_diameters = []
    for contour in contours:
        (x, y), radius = cv2.minEnclosingCircle(contour)
        center = (int(x), int(y))
        radius = int(radius)
        cv2.circle(image, center, radius, (255, 0, 0), 2)
        diameter = 2 * radius
        circle_diameters.append(diameter)
    return circle_diameters

def create_mask(image):
    """Create a mask based on green contours in the image."""
    lower_green = np.array([0, 255, 0])
    upper_green = np.array([0, 255, 0])
    mask = cv2.inRange(image, lower_green, upper_green)
    kernel = np.ones((5, 5), np.uint8)
    mask = cv2.dilate(mask, kernel, iterations=1)
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    mask_filled = np.zeros_like(mask)
    cv2.drawContours(mask_filled, contours, -1, (255), thickness=cv2.FILLED)
    return mask_filled

def calculate_area(mask):
    """Calculate the area size of the masked region."""
    area_size = cv2.countNonZero(mask)
    return area_size

def calculate_gray_stats(image, mask):
    """Verilen maske içinde ortalama, minimum ve maksimum gri değerlerini hesaplar."""
    masked_image = cv2.bitwise_and(image, image, mask=mask)
    gray_values = masked_image[np.where(mask != 0)]

    if gray_values.size == 0:
        return None, None, None  # Veya başka bir varsayılan değer döndürebilirsiniz

    mean_gray = np.mean(gray_values)
    min_gray = np.min(gray_values)
    max_gray = np.max(gray_values)

    return mean_gray, min_gray, max_gray


def calculate_rgb_stats(image, mask):
    """Calculate mean RGB values within the mask."""
    masked_rgb = cv2.bitwise_and(image, image, mask=mask)
    mean_rgb = cv2.mean(masked_rgb, mask=mask)
    return mean_rgb

def calculate_similarity(new_features, comparison_df):
    """Calculate similarity between new data and each row in comparison_df."""
    similarity_results = []

    for index, row in comparison_df.iterrows():
        sim = {}
        for feature in new_features.keys():
            if feature != "Görsel Adı" and feature in row and pd.notna(row[feature]):
                if isinstance(new_features[feature], (int, float)) and isinstance(row[feature], (int, float)):
                    sim[feature] = 1 - abs(new_features[feature] - row[feature]) / max(new_features[feature], row[feature])
                else:
                    sim[feature] = 1 if new_features[feature] == row[feature] else 0  # String or other types comparison
            else:
                sim[feature] = None  # or any other way to handle missing features
        similarity_results.append(sim)

    return similarity_results

def calculate_average_similarity(similarity_results):
    """Calculate the average similarity for each feature."""
    avg_similarity = {}
    feature_counts = {}

    for sim in similarity_results:
        for feature, value in sim.items():
            if value is not None:  # Ignore None values
                if feature in avg_similarity:
                    avg_similarity[feature] += value
                    feature_counts[feature] += 1
                else:
                    avg_similarity[feature] = value
                    feature_counts[feature] = 1

    for feature in avg_similarity.keys():
        avg_similarity[feature] /= feature_counts[feature]

    return avg_similarity


def process_excel(file_path):
    # Excel dosyasını yükleyin
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Satırları dolaşın ve karşılaştırmaları yapın
    for row in range(2, sheet.max_row + 1):  # 2. satırdan başlıyoruz çünkü 1. satır genellikle başlıklar için kullanılır
        x = 0
        if sheet[f'B{row}'].value < sheet[f'C{row}'].value:
            x += 2
        if sheet[f'D{row}'].value < sheet[f'E{row}'].value:
            x += 1
        if sheet[f'F{row}'].value > sheet[f'G{row}'].value:
            x += 1
        if sheet[f'H{row}'].value < sheet[f'I{row}'].value:
            x += 1
        if sheet[f'J{row}'].value > sheet[f'K{row}'].value:
            x += 1
        if sheet[f'L{row}'].value < sheet[f'M{row}'].value:
            x += 1
        if sheet[f'N{row}'].value < sheet[f'O{row}'].value:
            x += 1
       # if sheet[f'P{row}'].value < sheet[f'Q{row}'].value:
        #    x += 1

        # L sütununa sonucu yazın
        sheet[f'R{row}'] = x
        if x > 6:
            print("Görselin benzerlik derecesi:", x, "Kanser olma ihtimali yüksektir.")
        else:
            print("Görselin benzerlik derecesi:", x, "Kanser olma ihtimali düşüktür.")

    # Değişiklikleri kaydedin
    workbook.save(file_path)

if __name__ == '__main__':
    # Resim yolu ve adını belirleme
    # image_path = "./dataset/train/train/not_skin_cancer/not_skin_cancer_56.jpg"
    image_path = "./dataset/test/test/skin_cancer/skin_cancer_120.jpg"
    image_name = os.path.basename(image_path)
    white_path = "./dataset/test/beyaz.jpg"

    # Görüntüyü yükleme ve işleme
    origin = cv2.imread(image_path)
    white_img = cv2.imread(white_path)
    image = cv2.imread(image_path)
    image = cv2.resize(image, None, fx=3, fy=3, interpolation=cv2.INTER_LINEAR)
    image1 = gri(image)
    image2 = cv2.equalizeHist(image1)
    image3 = iso(image1)
    image4 = apply_median_filter(image3, 77)
    contours = find_contours(image4)
    draw_contours(image, contours)
    diameters = draw_circles(image, contours)
    white_path = draw_contours(white_img, contours)

    # Maske oluşturma
    mask = create_mask(image)

    # Alan boyutunu hesaplama
    area_size = calculate_area(mask)

    # Gri tonlama istatistiklerini hesaplama
    mean_gray, min_gray, max_gray = calculate_gray_stats(image1, mask)


    # RGB istatistiklerini hesaplama
    mean_rgb = calculate_rgb_stats(image, mask)

    # Yeni görselin özelliklerini bir sözlükte toplama
    new_data = {
        "Görsel Adı": image_name,
        "Alan Boyutu (piksel)": area_size,
        "Ortalama Gri Değer": mean_gray,
        "Minimum Gri Değer": min_gray,
        "Maksimum Gri Değer": max_gray,
        "Ortalama R Değeri": mean_rgb[2],
        "Ortalama G Değeri": mean_rgb[1],
        "Ortalama B Değeri": mean_rgb[0],
        "Sınır Sayısı": len(diameters)
    }

    print(new_data)

    # Sonuçları masaüstüne kaydetme
    desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    save_path = os.path.join(desktop_path, "white_img_result.jpg")
    cv2.imwrite(save_path, white_img)

    # Görsel özellikler ve not_cancer dosyalarını yükleme
    cancer = pd.read_excel("cancer.xlsx")
    not_cancer = pd.read_excel("not_cancer.xlsx")

    # Benzerlik hesaplamaları
    similarities_cancer = calculate_similarity(new_data, cancer)
    similarities_not_cancer = calculate_similarity(new_data, not_cancer)

    # Ortalama benzerlik sonuçlarını hesaplama
    avg_similarity_cancer = calculate_average_similarity(similarities_cancer)
    avg_similarity_not_cancer = calculate_average_similarity(similarities_not_cancer)

    # Sonuçları birleştirme
    combined_avg_similarity = {"Görsel Adı": image_name}
    for feature in new_data.keys():
        if feature in avg_similarity_cancer:
            combined_avg_similarity[f"{feature}1"] = avg_similarity_cancer[feature]
        if feature in avg_similarity_not_cancer:
            combined_avg_similarity[f"{feature}2"] = avg_similarity_not_cancer[feature]

    # Benzerlik sonuçlarını DataFrame'e dönüştürme
    similarity_df = pd.DataFrame([combined_avg_similarity])

    # Sonuçları Excel dosyasına kaydetme
    result_file_name = "sonuc.xlsx"
    if os.path.exists(result_file_name):
        with pd.ExcelWriter(result_file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            similarity_df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)


    else:
        similarity_df.to_excel(result_file_name, index=False)

    process_excel('sonuc.xlsx')

    # Görüntüleri aynı anda açma
    images = [origin, image1, image2, image3, image, image4, white_img]
    window_names = ["ORIGIN", "GRI", "HISTOGRAM", "ISO", "SINIRLI", "MEDIAN", "WHITE"]

    for window_name, img in zip(window_names, images):
        image_show(img, window_name)

    cv2.waitKey(0)
    cv2.destroyAllWindows()